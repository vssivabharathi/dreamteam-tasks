
import openpyxl
import sys

def multiplication(number):
    wb = openpyxl.Workbook()
    
    
    #Creating the workseet
    sheet = wb.active
    sheet.title = '{}x{} multiplication table'.format(number, number)

    #creating the row, column 1
    for i in range(number):
        sheet.cell(row=i+1, column=1).value=i
    
    # Creating the column at row 1
    for i in range(number):
        sheet.cell(column=i+1, row=1).value=i

    #for loop for the mutiplication 
    for row in range  (1, number+1):  
        for col in range (1, number+1):
            sheet.cell(row=row+1, column=col+1).value=row*col
    
    # saving the table 
    wb.save('mutiplication.xlsx')

# Check if the correct number of arguments is passed
if len(sys.argv) != 2:
    print("Usage: python script_name.py <number>")
    sys.exit(1)

try:
    # Get the number from command-line arguments
    number = int(sys.argv[1])
    multiplication(number)
    print(f"Multiplication table for {number} has been saved as 'multiplication_{number}.xlsx'")
except ValueError:
    print("Please enter a valid integer.")
    sys.exit(1)