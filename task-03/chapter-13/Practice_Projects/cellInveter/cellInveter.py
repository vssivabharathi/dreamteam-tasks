#inveting the cell row to column and column to row.
import openpyxl


def invertCells(filename):
  
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    newSheet = wb.create_sheet(index=0, title='inverted')

    for rowObj in sheet.rows:
        for cellObj in rowObj:
            col = cellObj.column
            row = cellObj.row

            newSheet.cell(row=col, column=row).value = cellObj.value

    wb.save('result_'+filename)


if __name__ == "__main__":
    invertCells('example.xlsx')
