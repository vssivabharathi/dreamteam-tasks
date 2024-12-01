import openpyxl
import sys

def insert_blank_rows(start_row, num_blank_rows, file_path): 
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    rows = tuple(sheet.rows)

    for row in rows[::-1]:
        for cell in row:
            col = cell.column
            row_num = cell.row

            if row_num >= start_row and row_num < start_row + num_blank_rows:
                sheet.cell(row=row_num + num_blank_rows, column=col).value = cell.value
                sheet.cell(row=row_num, column=col).value = ''
            elif row_num >= start_row + num_blank_rows:
                sheet.cell(row=row_num + num_blank_rows, column=col).value = cell.value

    
    wb.save('result_' + file_path)

if __name__ == "__main__": 
    num_args = len(sys.argv)
    if num_args < 4:
        print("Usage: python blankRowInserter.py <start_row> <num_blank_rows> <file_path>")
    else: 
        start_row = int(sys.argv[1])  
        num_blank_rows = int(sys.argv[2])  
        file_path = sys.argv[3] 
        insert_blank_rows(start_row, num_blank_rows, file_path)
